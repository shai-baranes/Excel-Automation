from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout

# Load workbook and sheet
wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

# Define data range for the chart
min_column = sheet.min_column  # 1
max_column = sheet.max_column  # 7
min_row = sheet.min_row        # 5
max_row = sheet.max_row        # 7

# Create a Bar Chart
chart = BarChart()

# Define data and categories
data = Reference(sheet, min_col=min_column + 1, max_col=max_column, min_row=min_row, max_row=max_row)  # Data includes headers
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row + 1, max_row=max_row)  # Categories (Female/Male)

# Add data and categories to the chart
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

# Chart title and style
chart.title = "Sales by Product Line"
chart.style = 10

# Adjust chart size (width and height in centimeters)
chart.width = 20  # Set width of the chart (in cm)
chart.height = 15  # Set height of the chart (in cm)


# Optional: Adjust overlap for better display
chart.overlap = 30

# Ensure axes are visible
chart.x_axis.delete = False
chart.y_axis.delete = False

# chart.legend.position = "b"  # Position legend at the bottom of the chart (nice potential tweak to avoid re-sizing...)

# Adjust plot area layout using ManualLayout
chart.layout = Layout(
    manualLayout=ManualLayout(
        x=-0.1,   # X position (proportion of container width)
        y=0.1,   # Y position (proportion of container height)
        w=0.6,   # Width (proportion of container width)
        # w=0.8,   # Width (proportion of container width)
        h=0.6    # Height (proportion of container height)
    )
)

# Add the chart to the worksheet at a specific position
sheet.add_chart(chart, "B12")

# Save the workbook with the updated chart layout
wb.save('chart_with_adjusted_plot_area.xlsx')
