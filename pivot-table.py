import pandas as pd

month = "January"

df = pd.read_excel('supermarket_sales.xlsx')

df = df[['Gender', 'Product line','Total']]


pivot_table1 = df.pivot_table(index = 'Gender', columns = 'Product line', values = 'Total', aggfunc = 'sum').round(0) # check how to add more values here
pivot_table2 = df.pivot_table(index = 'Product line', columns = 'Gender', values = 'Total', aggfunc = 'sum').round(0) # check how to add more values here

# pivot_table.columns --> ['Electronic accessories', 'Fashion accessories', 'Food and beverages', 'Health and beauty', 'Home and lifestyle', 'Sports and travel']

# pivot_table --> 
# Product line  Electronic accessories  ...  Sports and travel
# Gender                                ...                   
# Female                    27102.0225  ...         28574.7210
# Male                      27235.5090  ...         26548.1055


## note that we cannot add 2 sheets without using the Excel Writer (as I also did in EPD for multiple sheets per Excel file)
# pivot_table1.to_excel('pivot_table.xlsx', 'Report1', startrow=4) # other are optional params
# pivot_table2.to_excel('pivot_table.xlsx', 'Report2', startrow=4) # other are optional params

# Use ExcelWriter to write multiple sheets in a single file
with pd.ExcelWriter(f'./Results/report_{month}.xlsx') as writer:
    pivot_table1.to_excel(writer, sheet_name='Report1', startrow=4)  # Write first sheet
    pivot_table2.to_excel(writer, sheet_name='Report2', startrow=4)  # Write second sheet


# ---------------------------------
#    adding the charts
# ---------------------------------


from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, BarChart3D
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font



wb = load_workbook(f'./Results/report_{month}.xlsx')

sheet1 = wb['Report1']
sheet2 = wb['Report2']

# # now to select the active rows and columns ('pivot table' location)
# min_column = wb.active.min_column # 1
# max_column = wb.active.max_column # 7
# min_row = wb.active.min_row # 5
# max_row = wb.active.max_row # 7

# now to select the active rows and columns ('pivot table' location - for sheet#1)
sheet1_min_column = sheet1.min_column # 1
sheet1_max_column = sheet1.max_column # 7
sheet1_min_row = sheet1.min_row # 5
sheet1_max_row = sheet1.max_row # 7

# now to select the active rows and columns ('pivot table' location - for sheet#2)
sheet2_min_column = sheet2.min_column # 1
sheet2_max_column = sheet2.max_column # 7
sheet2_min_row = sheet2.min_row # 5
sheet2_max_row = sheet2.max_row # 7



chart1 = BarChart() # TBD try also barchart from the pandas embedded charts
chart2 = BarChart() # TBD try also barchart from the pandas embedded charts

data1 = Reference(sheet1, min_col = sheet1_min_column + 1, max_col = sheet1_max_column, min_row = sheet1_min_row, max_row = sheet1_max_row) # data include the headers (but not categories)
categories1 = Reference(sheet1, min_col = sheet1_min_column, max_col = sheet1_min_column, min_row = sheet1_min_row + 1, max_row = sheet1_max_row) # only 'Female / Male'

chart1.add_data(data = data1, titles_from_data = True)
chart1.set_categories(categories1)

data2 = Reference(sheet2, min_col = sheet2_min_column + 1, max_col = sheet2_max_column, min_row = sheet2_min_row, max_row = sheet2_max_row) # data include the headers (but not categories)
categories2 = Reference(sheet2, min_col = sheet2_min_column, max_col = sheet2_min_column, min_row = sheet2_min_row + 1, max_row = sheet2_max_row) # only 'Female / Male'

chart2.add_data(data = data2, titles_from_data = True)
chart2.set_categories(categories2)



chart1.title = "Sales by Product Line"
chart1.style = 10

chart2.title = "Sales by Product Line"
chart2.style = 10

# Adjust chart size (width and height in centimeters)
chart1.width = 20  # Set width of the chart (in cm)
chart1.height = 15  # Set height of the chart (in cm)

# Adjust chart size (width and height in centimeters)
chart2.width = 20  # Set width of the chart (in cm)
chart2.height = 15  # Set height of the chart (in cm)


# ----
# new
# chart.grouping = "percentStacked"
chart1.overlap = 30 # nice disaplay :)
chart2.overlap = 30 # nice disaplay :)
# ----


# to enable the axes view:
chart1.x_axis.delete = False
chart1.y_axis.delete = False

# to enable the axes view:
chart2.x_axis.delete = False
chart2.y_axis.delete = False

# Adjust plot area layout using ManualLayout (sheet#1)
chart1.layout = Layout(
    manualLayout=ManualLayout(
        x=-0.1,   # X position (proportion of container width)
        y=0.1,   # Y position (proportion of container height)
        w=0.6,   # Width (proportion of container width)
        # w=0.8,   # Width (proportion of container width)
        h=0.6    # Height (proportion of container height)
    )
)

# Adjust plot area layout using ManualLayout (sheet#2)
chart2.layout = Layout(
    manualLayout=ManualLayout(
        x=-0.1,   # X position (proportion of container width)
        y=0.1,   # Y position (proportion of container height)
        w=0.6,   # Width (proportion of container width)
        # w=0.8,   # Width (proportion of container width)
        h=0.6    # Height (proportion of container height)
    )
)

sheet1.add_chart(chart1, "B12")
sheet2.add_chart(chart2, "B15")


# # Adding SUM to the tables using python automation (per active column)
# # herein a simple example to be enhanced
# sheet1['B8'] = '=SUM(B6:B7)' # note that the formula was taken directly from the excel
# sheet1['B8'].style = Currency # note that the formula was taken directly from the excel

for i in range(sheet1_min_column+1, sheet1_max_column+1):
	letter = get_column_letter(i)
	sheet1[f'{letter}{sheet1_max_row+1}'] = f'=SUM({letter}{sheet1_min_row+1}:{letter}{sheet1_max_row})'
	sheet1[f'{letter}{sheet1_max_row+1}'].style = 'Currency'


for i in range(sheet2_min_column+1, sheet2_max_column+1):
	letter = get_column_letter(i)
	sheet2[f'{letter}{sheet2_max_row+1}'] = f'=SUM({letter}{sheet2_min_row+1}:{letter}{sheet2_max_row})'
	sheet2[f'{letter}{sheet2_max_row+1}'].style = 'Currency'


# Adding Title & Sub-title (you can also add tested versions, date, tester name etc... after the report creation)
sheet1['A1'] = 'Sales Report'
sheet1['A2'] = month
# sheet1['A2'] = 'January'
sheet1['A1'].font = Font('Arial', bold=True, size=20)
sheet1['A2'].font = Font('Arial', bold=True, size=15)




wb.save(f'./Results/report_{month}.xlsx')
# wb.save('./Results/pivot_table.xlsx')
# wb.save('chart.xlsx') # I sometimes write the charts (&tables) into another excel file for debug needs and for having the first doc const. open to the side















